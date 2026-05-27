from pathlib import Path
from copy import copy
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter


SOURCE = Path("/Users/jamesmaxwell/Downloads/Copy of Redmond Profit Matrix .xlsx")
OUT_DIR = Path("/Users/jamesmaxwell/Documents/Codex/2026-05-14/files-mentioned-by-the-user-copy/outputs/amazon_sku_pnl")
OUT = OUT_DIR / "Redmond Amazon SKU P&L Scenario Model.xlsx"


def safe_num(value, default=0):
    if value is None:
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def short_title(title):
    if not title:
        return ""
    text = str(title).replace("REDMOND ", "").replace("Redmond ", "")
    return text[:75]


def style_range(ws, cell_range, fill=None, font=None, border=None, alignment=None):
    for row in ws[cell_range]:
        for cell in row:
            if fill:
                cell.fill = fill
            if font:
                cell.font = font
            if border:
                cell.border = border
            if alignment:
                cell.alignment = alignment


def set_col_widths(ws, widths):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def build():
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    src_values = load_workbook(SOURCE, data_only=True)
    src_formulas = load_workbook(SOURCE, data_only=False)
    pnl_src = src_values["Profit Matrix Per SKU Last 30 d"]

    rows = []
    for r in range(3, pnl_src.max_row + 1):
        asin = pnl_src.cell(r, 1).value
        title = pnl_src.cell(r, 3).value
        if not asin or not title:
            continue
        sales = safe_num(pnl_src.cell(r, 4).value)
        units = safe_num(pnl_src.cell(r, 5).value)
        if sales == 0 and units == 0:
            continue
        sku = pnl_src.cell(r, 2).value or ""
        if isinstance(sku, str) and sku.strip().upper() == "#N/A":
            sku = ""
        rows.append(
            {
                "asin": asin,
                "sku": sku,
                "title": short_title(title),
                "sales": sales,
                "units": units,
                "amazon_fees": safe_num(pnl_src.cell(r, 10).value),
                "ad_spend": safe_num(pnl_src.cell(r, 12).value),
                "cogs": safe_num(pnl_src.cell(r, 14).value),
                "profit": safe_num(pnl_src.cell(r, 16).value),
                "margin": safe_num(pnl_src.cell(r, 17).value),
            }
        )

    rows.sort(key=lambda x: x["sales"], reverse=True)

    wb = Workbook()
    wb.remove(wb.active)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"

    dashboard = wb.create_sheet("Dashboard")
    assumptions = wb.create_sheet("Assumptions")
    pnl = wb.create_sheet("SKU P&L")
    sensitivity = wb.create_sheet("Sensitivity")
    checks = wb.create_sheet("Checks")

    # Palette
    navy = "183B56"
    teal = "168A8A"
    coral = "D96B5F"
    gold = "F3C969"
    ink = "24313A"
    mist = "EEF5F4"
    soft_blue = "EAF1F7"
    soft_yellow = "FFF4CC"
    pale_green = "E7F4EA"
    pale_red = "FCE8E6"
    grid = "D9E2E1"
    white = "FFFFFF"

    thin = Side(style="thin", color=grid)
    medium = Side(style="medium", color="AFC1BF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor=navy)
    section_fill = PatternFill("solid", fgColor=teal)
    input_fill = PatternFill("solid", fgColor=soft_yellow)
    output_fill = PatternFill("solid", fgColor=soft_blue)
    body_fill = PatternFill("solid", fgColor=white)
    formula_font = Font(name="Aptos", color="000000", size=10)
    input_font = Font(name="Aptos", color="0000FF", size=10)
    title_font = Font(name="Aptos Display", color=white, bold=True, size=20)
    subtitle_font = Font(name="Aptos", color="DDEDEA", size=10)
    header_font = Font(name="Aptos", color=white, bold=True, size=10)
    label_font = Font(name="Aptos", color=ink, bold=True, size=10)
    body_font = Font(name="Aptos", color=ink, size=10)

    # Assumptions
    assumptions.sheet_view.showGridLines = False
    assumptions.merge_cells("A1:F1")
    assumptions["A1"] = "Amazon SKU P&L Scenario Assumptions"
    assumptions["A1"].fill = header_fill
    assumptions["A1"].font = title_font
    assumptions["A1"].alignment = Alignment(horizontal="left", vertical="center")
    assumptions.row_dimensions[1].height = 30
    assumptions.merge_cells("A2:F2")
    assumptions["A2"] = "Blue/yellow cells are editable. Per-SKU overrides on the SKU P&L sheet take precedence over these global defaults."
    assumptions["A2"].fill = PatternFill("solid", fgColor=navy)
    assumptions["A2"].font = subtitle_font

    assumption_rows = [
        ("Default target TACOS", 0.09, "Ad spend as % of scenario revenue"),
        ("Default coupon % of sales", 0.00, "Coupon cost as % of scenario revenue"),
        ("Default coupon $ / unit", 0.00, "Dollar coupon or promo cost per unit"),
        ("Default sales lift", 0.00, "Expected unit/revenue lift from coupons or pricing"),
        ("Profit margin warning", 0.25, "Rows below this scenario margin are flagged"),
        ("Currency", "USD", "Model is built in dollars"),
    ]
    for col, value in enumerate(["Input", "Value", "Notes", "", "Formula color key", "Meaning"], start=1):
        assumptions.cell(3, col, value)
    for row_idx, row in enumerate(assumption_rows, start=4):
        for col_idx, value in enumerate(list(row) + ["", "", ""], start=1):
            assumptions.cell(row_idx, col_idx, value)
    assumptions["E4"] = "Blue text / yellow fill"
    assumptions["F4"] = "Editable inputs"
    assumptions["E5"] = "Black text"
    assumptions["F5"] = "Formulas and calculated outputs"
    assumptions["E6"] = "Green text"
    assumptions["F6"] = "Links to other sheets"

    for row in range(4, 4 + len(assumption_rows)):
        assumptions[f"A{row}"].font = label_font
        assumptions[f"B{row}"].fill = input_fill
        assumptions[f"B{row}"].font = input_font
        assumptions[f"C{row}"].font = body_font
        assumptions[f"B{row}"].number_format = "0.0%" if row in [4, 5, 7, 8] else "$0.00"
        if row == 9:
            assumptions[f"B{row}"].number_format = "@"
    style_range(assumptions, "A4:F10", border=border, alignment=Alignment(vertical="center"))
    style_range(assumptions, "A3:F3", fill=section_fill, font=header_font, border=border)
    assumptions.column_dimensions["A"].width = 28
    assumptions.column_dimensions["B"].width = 18
    assumptions.column_dimensions["C"].width = 50
    assumptions.column_dimensions["E"].width = 24
    assumptions.column_dimensions["F"].width = 30
    assumptions.freeze_panes = "A4"

    percent_validation = DataValidation(type="decimal", operator="between", formula1="0", formula2="1", allow_blank=True)
    percent_validation.error = "Enter a percentage between 0% and 100%."
    assumptions.add_data_validation(percent_validation)
    for cell in ["B4", "B5", "B7", "B8"]:
        percent_validation.add(assumptions[cell])

    # SKU P&L
    pnl.sheet_view.showGridLines = False
    pnl.merge_cells("A1:AC1")
    pnl["A1"] = "Per SKU P&L and Scenario Model"
    pnl["A1"].fill = header_fill
    pnl["A1"].font = title_font
    pnl["A1"].alignment = Alignment(horizontal="left")
    pnl.merge_cells("A2:AC2")
    pnl["A2"] = "Change the yellow override columns to test coupons, TACOS goals, ad spend, and sales lift at SKU level."
    pnl["A2"].fill = PatternFill("solid", fgColor=navy)
    pnl["A2"].font = subtitle_font

    headers = [
        "ASIN", "SKU", "Product", "Current Sales", "Units", "ASP", "Current Ad Spend", "Current TACOS",
        "Current Amazon Fees", "Fee %", "Current COGS", "COGS %", "Current Profit", "Current Margin",
        "Target TACOS Override", "Coupon % Override", "Coupon $/Unit Override", "Sales Lift Override",
        "Scenario Sales", "Scenario Units", "Scenario Coupon Cost", "Scenario Ad Spend",
        "Scenario Amazon Fees", "Scenario COGS", "Scenario Profit", "Scenario Margin",
        "Profit Delta", "Margin Delta", "Break-even TACOS"
    ]
    for col, header in enumerate(headers, start=1):
        pnl.cell(3, col, header)

    start_row = 4
    for idx, item in enumerate(rows, start=start_row):
        pnl.cell(idx, 1, item["asin"])
        pnl.cell(idx, 2, item["sku"])
        pnl.cell(idx, 3, item["title"])
        pnl.cell(idx, 4, item["sales"])
        pnl.cell(idx, 5, item["units"])
        pnl.cell(idx, 6, f'=IFERROR(D{idx}/E{idx},0)')
        pnl.cell(idx, 7, item["ad_spend"])
        pnl.cell(idx, 8, f'=IFERROR(G{idx}/D{idx},0)')
        pnl.cell(idx, 9, item["amazon_fees"])
        pnl.cell(idx, 10, f'=IFERROR(I{idx}/D{idx},0)')
        pnl.cell(idx, 11, item["cogs"])
        pnl.cell(idx, 12, f'=IFERROR(K{idx}/D{idx},0)')
        pnl.cell(idx, 13, item["profit"])
        pnl.cell(idx, 14, f'=IFERROR(M{idx}/D{idx},0)')
        pnl.cell(idx, 15, None)
        pnl.cell(idx, 16, None)
        pnl.cell(idx, 17, None)
        pnl.cell(idx, 18, None)
        pnl.cell(idx, 19, f'=D{idx}*(1+IF(R{idx}="",Assumptions!$B$7,R{idx}))')
        pnl.cell(idx, 20, f'=E{idx}*(1+IF(R{idx}="",Assumptions!$B$7,R{idx}))')
        pnl.cell(idx, 21, f'=S{idx}*IF(P{idx}="",Assumptions!$B$5,P{idx})+T{idx}*IF(Q{idx}="",Assumptions!$B$6,Q{idx})')
        pnl.cell(idx, 22, f'=S{idx}*IF(O{idx}="",Assumptions!$B$4,O{idx})')
        pnl.cell(idx, 23, f'=S{idx}*J{idx}')
        pnl.cell(idx, 24, f'=S{idx}*L{idx}')
        pnl.cell(idx, 25, f'=S{idx}-U{idx}-V{idx}-W{idx}-X{idx}')
        pnl.cell(idx, 26, f'=IFERROR(Y{idx}/S{idx},0)')
        pnl.cell(idx, 27, f'=Y{idx}-M{idx}')
        pnl.cell(idx, 28, f'=Z{idx}-N{idx}')
        pnl.cell(idx, 29, f'=MAX(0,IFERROR((S{idx}-U{idx}-W{idx}-X{idx})/S{idx},0))')

    end_row = start_row + len(rows) - 1
    table_ref = f"A3:AC{end_row}"
    tab = Table(displayName="tblSkuPnl", ref=table_ref)
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    pnl.add_table(tab)
    pnl.freeze_panes = "D4"
    pnl.auto_filter.ref = table_ref

    for cell in pnl[3]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(top=medium, bottom=medium)
    for row in pnl.iter_rows(min_row=start_row, max_row=end_row, min_col=1, max_col=29):
        for cell in row:
            cell.font = formula_font if cell.column >= 6 else body_font
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            if cell.column in [15, 16, 17, 18]:
                cell.fill = input_fill
                cell.font = input_font
            elif cell.column >= 19:
                cell.fill = output_fill
            else:
                cell.fill = body_fill

    currency_cols = [4, 6, 7, 9, 11, 13, 17, 19, 21, 22, 23, 24, 25, 27]
    percent_cols = [8, 10, 12, 14, 15, 16, 18, 26, 28, 29]
    count_cols = [5, 20]
    for col in currency_cols:
        for row in range(start_row, end_row + 1):
            pnl.cell(row, col).number_format = '$#,##0;[Red]($#,##0);-'
    for col in percent_cols:
        for row in range(start_row, end_row + 1):
            pnl.cell(row, col).number_format = '0.0%;[Red](0.0%);-'
    for col in count_cols:
        for row in range(start_row, end_row + 1):
            pnl.cell(row, col).number_format = '#,##0'
    for col in range(1, 30):
        pnl.cell(3, col).font = header_font

    set_col_widths(pnl, {
        "A": 14, "B": 22, "C": 46, "D": 14, "E": 11, "F": 10, "G": 14, "H": 11,
        "I": 15, "J": 10, "K": 13, "L": 10, "M": 14, "N": 12, "O": 13, "P": 13,
        "Q": 13, "R": 12, "S": 14, "T": 12, "U": 15, "V": 15, "W": 16,
        "X": 13, "Y": 15, "Z": 13, "AA": 13, "AB": 12, "AC": 13
    })
    pnl.row_dimensions[3].height = 42

    pnl.conditional_formatting.add(f"Z{start_row}:Z{end_row}", ColorScaleRule(start_type="num", start_value=0, start_color=coral, mid_type="percentile", mid_value=50, mid_color="FFFFFF", end_type="max", end_color=teal))
    pnl.conditional_formatting.add(f"Y{start_row}:Y{end_row}", CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill("solid", fgColor=pale_red)))
    pnl.conditional_formatting.add(f"Z{start_row}:Z{end_row}", CellIsRule(operator="greaterThan", formula=["0"], fill=PatternFill("solid", fgColor=pale_green)))

    # Dashboard
    dashboard.sheet_view.showGridLines = False
    dashboard.merge_cells("A1:L1")
    dashboard["A1"] = "Redmond Amazon SKU P&L Dashboard"
    dashboard["A1"].fill = header_fill
    dashboard["A1"].font = title_font
    dashboard["A1"].alignment = Alignment(horizontal="left")
    dashboard.merge_cells("A2:L2")
    dashboard["A2"] = "Last-30-day baseline with live scenario outputs driven by coupon, TACOS, ad spend, and sales-lift assumptions."
    dashboard["A2"].fill = PatternFill("solid", fgColor=navy)
    dashboard["A2"].font = subtitle_font
    dashboard.row_dimensions[1].height = 31

    kpis = [
        ("Current Sales", f"=SUM('SKU P&L'!D{start_row}:D{end_row})"),
        ("Scenario Sales", f"=SUM('SKU P&L'!S{start_row}:S{end_row})"),
        ("Current Profit", f"=SUM('SKU P&L'!M{start_row}:M{end_row})"),
        ("Scenario Profit", f"=SUM('SKU P&L'!Y{start_row}:Y{end_row})"),
        ("Profit Delta", f"=D6-B6"),
        ("Scenario Margin", f"=IFERROR(D6/B5,0)"),
        ("Scenario Ad Spend", f"=SUM('SKU P&L'!V{start_row}:V{end_row})"),
        ("Scenario Coupon Cost", f"=SUM('SKU P&L'!U{start_row}:U{end_row})"),
    ]
    positions = ["A4", "C4", "A7", "C7", "E4", "G4", "E7", "G7"]
    for (label, formula), pos in zip(kpis, positions):
        col = dashboard[pos].column
        row = dashboard[pos].row
        dashboard.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
        dashboard.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 1)
        dashboard.cell(row, col, label)
        dashboard.cell(row + 1, col, formula)
        style_range(dashboard, f"{get_column_letter(col)}{row}:{get_column_letter(col+1)}{row+1}", border=border)
        dashboard.cell(row, col).fill = PatternFill("solid", fgColor=mist)
        dashboard.cell(row, col).font = label_font
        dashboard.cell(row + 1, col).fill = PatternFill("solid", fgColor=white)
        dashboard.cell(row + 1, col).font = Font(name="Aptos Display", color=ink, bold=True, size=16)
        dashboard.cell(row + 1, col).number_format = '0.0%' if "Margin" in label else '$#,##0;[Red]($#,##0);-'

    dashboard["A11"] = "Top 10 Scenario Profit Drivers"
    dashboard["A11"].fill = section_fill
    dashboard["A11"].font = header_font
    dashboard.merge_cells("A11:F11")
    dash_headers = ["Product", "Scenario Sales", "Scenario Profit", "Margin", "Profit Delta", "Target TACOS"]
    for c, h in enumerate(dash_headers, start=1):
        dashboard.cell(12, c, h)
        dashboard.cell(12, c).fill = header_fill
        dashboard.cell(12, c).font = header_font
        dashboard.cell(12, c).alignment = Alignment(horizontal="center", wrap_text=True)
    top = sorted(range(start_row, end_row + 1), key=lambda r: rows[r - start_row]["sales"], reverse=True)[:10]
    for out_row, pnl_row in enumerate(top, start=13):
        dashboard.cell(out_row, 1, f"='SKU P&L'!C{pnl_row}")
        dashboard.cell(out_row, 2, f"='SKU P&L'!S{pnl_row}")
        dashboard.cell(out_row, 3, f"='SKU P&L'!Y{pnl_row}")
        dashboard.cell(out_row, 4, f"='SKU P&L'!Z{pnl_row}")
        dashboard.cell(out_row, 5, f"='SKU P&L'!AA{pnl_row}")
        dashboard.cell(out_row, 6, f"=IF('SKU P&L'!O{pnl_row}=\"\",Assumptions!$B$4,'SKU P&L'!O{pnl_row})")
    for row in range(12, 23):
        for col in range(1, 7):
            dashboard.cell(row, col).border = border
            dashboard.cell(row, col).alignment = Alignment(vertical="center")
        for col in [2, 3, 5]:
            dashboard.cell(row, col).number_format = '$#,##0;[Red]($#,##0);-'
        for col in [4, 6]:
            dashboard.cell(row, col).number_format = '0.0%;[Red](0.0%);-'
    dashboard.conditional_formatting.add("E13:E22", ColorScaleRule(start_type="min", start_color=coral, mid_type="percentile", mid_value=50, mid_color="FFFFFF", end_type="max", end_color=teal))

    # Chart helper area and charts
    dashboard["H11"] = "Chart Data"
    dashboard["H11"].fill = section_fill
    dashboard["H11"].font = header_font
    chart_headers = ["Product", "Current Profit", "Scenario Profit"]
    for c, h in enumerate(chart_headers, start=8):
        dashboard.cell(12, c, h)
        dashboard.cell(12, c).fill = header_fill
        dashboard.cell(12, c).font = header_font
    for i, pnl_row in enumerate(top[:8], start=13):
        dashboard.cell(i, 8, f"=LEFT('SKU P&L'!C{pnl_row},28)")
        dashboard.cell(i, 9, f"='SKU P&L'!M{pnl_row}")
        dashboard.cell(i, 10, f"='SKU P&L'!Y{pnl_row}")
    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = "Current vs Scenario Profit"
    chart.y_axis.title = "SKU"
    chart.x_axis.title = "Profit ($)"
    chart.height = 8.0
    chart.width = 15.5
    chart.add_data(Reference(dashboard, min_col=9, max_col=10, min_row=12, max_row=20), titles_from_data=True)
    chart.set_categories(Reference(dashboard, min_col=8, min_row=13, max_row=20))
    chart.legend.position = "b"
    dashboard.add_chart(chart, "H23")

    set_col_widths(dashboard, {"A": 28, "B": 13, "C": 16, "D": 13, "E": 16, "F": 12, "G": 16, "H": 30, "I": 14, "J": 14, "K": 2, "L": 2})
    dashboard.freeze_panes = "A12"

    # Sensitivity
    sensitivity.sheet_view.showGridLines = False
    sensitivity.merge_cells("A1:J1")
    sensitivity["A1"] = "Portfolio Profit Sensitivity"
    sensitivity["A1"].fill = header_fill
    sensitivity["A1"].font = title_font
    sensitivity.merge_cells("A2:J2")
    sensitivity["A2"] = "Scenario profit at different TACOS goals and coupon rates. Values use current sales volume and current unit economics."
    sensitivity["A2"].fill = PatternFill("solid", fgColor=navy)
    sensitivity["A2"].font = subtitle_font
    tac_range = [0.05, 0.07, 0.09, 0.11, 0.13, 0.15]
    coupon_range = [0.00, 0.03, 0.05, 0.08, 0.10]
    sensitivity["A4"] = "Coupon % \\ TACOS"
    for c, tac in enumerate(tac_range, start=2):
        sensitivity.cell(4, c, tac)
        sensitivity.cell(4, c).number_format = "0.0%"
    for r, coupon in enumerate(coupon_range, start=5):
        sensitivity.cell(r, 1, coupon)
        sensitivity.cell(r, 1).number_format = "0.0%"
        for c, tac in enumerate(tac_range, start=2):
            sensitivity.cell(r, c, f"=SUM('SKU P&L'!S${start_row}:S${end_row})*(1-{tac}-{coupon})-SUM('SKU P&L'!W${start_row}:W${end_row})-SUM('SKU P&L'!X${start_row}:X${end_row})")
            sensitivity.cell(r, c).number_format = '$#,##0;[Red]($#,##0);-'
    style_range(sensitivity, "A4:G9", border=border)
    style_range(sensitivity, "A4:G4", fill=header_fill, font=header_font)
    style_range(sensitivity, "A5:A9", fill=section_fill, font=header_font)
    sensitivity.conditional_formatting.add("B5:G9", ColorScaleRule(start_type="min", start_color=coral, mid_type="percentile", mid_value=50, mid_color="FFFFFF", end_type="max", end_color=teal))
    set_col_widths(sensitivity, {"A": 18, "B": 14, "C": 14, "D": 14, "E": 14, "F": 14, "G": 14})

    line = LineChart()
    line.title = "Profit by Coupon Rate at TACOS Goals"
    line.y_axis.title = "Profit ($)"
    line.x_axis.title = "Coupon Rate"
    line.height = 8
    line.width = 15
    line.add_data(Reference(sensitivity, min_col=2, max_col=7, min_row=4, max_row=9), titles_from_data=True)
    line.set_categories(Reference(sensitivity, min_col=1, min_row=5, max_row=9))
    line.legend.position = "b"
    sensitivity.add_chart(line, "A12")

    # Checks
    checks.sheet_view.showGridLines = False
    checks.merge_cells("A1:F1")
    checks["A1"] = "Model Checks"
    checks["A1"].fill = header_fill
    checks["A1"].font = title_font
    checks.append([])
    checks.append(["Check", "Actual", "Expected", "Difference", "Status", "Notes"])
    check_rows = [
        ("Current sales tie-out", f"=SUM('SKU P&L'!D{start_row}:D{end_row})", f"=SUM('Source - Current P&L'!D2:D{len(rows)+1})", "=B4-C4", '=IF(ABS(D4)<1,"OK","Review")', "SKU P&L should tie to source current P&L"),
        ("Current profit tie-out", f"=SUM('SKU P&L'!M{start_row}:M{end_row})", f"=SUM('Source - Current P&L'!I2:I{len(rows)+1})", "=B5-C5", '=IF(ABS(D5)<1,"OK","Review")', "Baseline profit should tie to source"),
        ("Scenario margin valid", f"=MIN('SKU P&L'!Z{start_row}:Z{end_row})", "=-1", "=B6-C6", '=IF(B6>=-1,"OK","Review")', "Margin floor sanity check"),
    ]
    for row in check_rows:
        checks.append(row)
    style_range(checks, "A3:F6", border=border)
    style_range(checks, "A3:F3", fill=section_fill, font=header_font)
    for row in range(4, 7):
        checks.cell(row, 2).number_format = '$#,##0;[Red]($#,##0);-' if row < 6 else "0.0%"
        checks.cell(row, 3).number_format = '$#,##0;[Red]($#,##0);-' if row < 6 else "0.0%"
        checks.cell(row, 4).number_format = '$#,##0;[Red]($#,##0);-' if row < 6 else "0.0%"
    checks.conditional_formatting.add("E4:E6", CellIsRule(operator="equal", formula=['"OK"'], fill=PatternFill("solid", fgColor=pale_green)))
    checks.conditional_formatting.add("E4:E6", CellIsRule(operator="notEqual", formula=['"OK"'], fill=PatternFill("solid", fgColor=pale_red)))
    set_col_widths(checks, {"A": 28, "B": 18, "C": 18, "D": 18, "E": 12, "F": 48})

    # Source current P&L data for audit/refresh.
    source_sheet = wb.create_sheet("Source - Current P&L")
    source_sheet.sheet_view.showGridLines = False
    source_sheet.append(["ASIN", "SKU", "Product", "Current Sales", "Units", "Amazon Fees", "Ad Spend", "COGS", "Profit", "Margin"])
    for item in rows:
        source_sheet.append([
            item["asin"], item["sku"], item["title"], item["sales"], item["units"], item["amazon_fees"],
            item["ad_spend"], item["cogs"], item["profit"], item["margin"]
        ])
    tab2 = Table(displayName="tblSourceCurrentPnl", ref=f"A1:J{len(rows)+1}")
    tab2.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showFirstColumn=False, showLastColumn=False, showRowStripes=True)
    source_sheet.add_table(tab2)
    for cell in source_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    for row in range(2, len(rows) + 2):
        for col in [4, 6, 7, 8, 9]:
            source_sheet.cell(row, col).number_format = '$#,##0;[Red]($#,##0);-'
        source_sheet.cell(row, 10).number_format = '0.0%;[Red](0.0%);-'
    set_col_widths(source_sheet, {"A": 14, "B": 22, "C": 48, "D": 14, "E": 11, "F": 14, "G": 14, "H": 12, "I": 14, "J": 12})
    source_sheet.freeze_panes = "A2"

    # Copy raw source tabs behind the model for auditability.
    for src_name in ["Last 30 day Sales", "Ad Spend Per SKU", "RL COGs", "FBA Fees"]:
        src_ws = src_formulas[src_name]
        dst = wb.create_sheet(f"Raw - {src_name[:24]}")
        for row in src_ws.iter_rows():
            for src_cell in row:
                dst_cell = dst.cell(src_cell.row, src_cell.column, src_cell.value)
                if src_cell.has_style:
                    dst_cell.font = copy(src_cell.font)
                    dst_cell.fill = copy(src_cell.fill)
                    dst_cell.border = copy(src_cell.border)
                    dst_cell.alignment = copy(src_cell.alignment)
                    dst_cell.number_format = src_cell.number_format
        for key, dim in src_ws.column_dimensions.items():
            dst.column_dimensions[key].width = dim.width
        dst.freeze_panes = "A2"

    # Workbook-wide page and alignment polish.
    for ws in wb.worksheets:
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_margins.left = 0.25
        ws.page_margins.right = 0.25
        ws.page_margins.top = 0.35
        ws.page_margins.bottom = 0.35
        for row in ws.iter_rows():
            for cell in row:
                if not cell.font:
                    cell.font = body_font

    wb.save(OUT)
    return OUT


if __name__ == "__main__":
    print(build())
